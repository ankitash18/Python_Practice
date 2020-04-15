import openpyxl


wb = openpyxl.load_workbook('C:\\Users\\AShrivastava\\Desktop\\study\\sparkrockthejvm\\python\\store.xlsx')

#wb is an iterable
print(wb.worksheets)

for sheet in wb:
    print(sheet.title)

sheet1 = wb['Products']

sheet1 = wb.active

b2_cell = sheet1['B2']

c2_cell = sheet1['A2']

print(b2_cell.value)
print(c2_cell.value)
print(c2_cell.row,c2_cell.column)

print(sheet1['A2'].parent)


cell_range = sheet1['A2:B13']

for month,price in cell_range:
    print(f'month :{month.value}...price..{price.value}')


print(f'shet dimesnin ..{sheet1.dimensions}')

for a,b in sheet1[sheet1.dimensions]:
    print(a.value,',',b.value)


sheet['B2'] = 1000

wb.save('C:\\Users\\AShrivastava\\Desktop\\study\\sparkrockthejvm\\python\\store.xlsx')

new_sales = ('JAN',2000)
sheet1.append(new_sales)

wb.save('C:\\Users\\AShrivastava\\Desktop\\study\\sparkrockthejvm\\python\\store.xlsx')

wb.create_sheet('Turnover')

wb.save('C:\\Users\\AShrivastava\\Desktop\\study\\sparkrockthejvm\\python\\store.xlsx')
#create new excel file

wb1 = openpyxl.Workbook()

s1 = wb1.active

sales = {2017:70000,2018:8000,2019:90000}

s1['A1'] = 'YEAR'
s1['B1'] = 'SALES'

for k,v in sales.items():
    s1.append((k,v))

wb1.save('C:\\Users\\AShrivastava\\Desktop\\study\\sparkrockthejvm\\python\\salesnew.xlsx')



